
import re

from openpyxl.utils import column_index_from_string, get_column_letter

from .tokens import applyTokenReplacement

def resolveLookups(wb, elements = [], unprocessedDefinitions = [], currentElement = {}):
    if len(unprocessedDefinitions) == 0:
        elements.append(currentElement)
    else:
        loopDefinition = unprocessedDefinitions[0]

        type = loopDefinition.get("type", "unknown").lower()
        if type not in ["loopsheets", "findrow", "findcolumn", "looprows", "loopcolumns"]:
            raise ValueError(f"Invalid loop type '{type}' in definition: {loopDefinition}")
        
        if "token" not in loopDefinition:
            raise ValueError(f"Missing 'token' in loop definition: {loopDefinition}")
        token = loopDefinition["token"]
        
        loopElements = []
        
        if type == "loopsheets":
            if "regex" not in loopDefinition:
                raise ValueError(f"Missing 'regex' in loop definition: {loopDefinition}")
            sheetRegex = loopDefinition["regex"]
            matchingSheets = [sheet.title for sheet in wb.worksheets if re.search(sheetRegex, sheet.title)]
            loopElements = matchingSheets

        elif type == "findrow" or type == "findcolumn":
            if "regex" not in loopDefinition:
                raise ValueError(f"Missing 'regex' in loop definition: {loopDefinition}")
            regex = loopDefinition["regex"]

            if "sheet" not in loopDefinition:
                raise ValueError(f"Missing 'sheet' in loop definition: {loopDefinition}")
            sheet = applyTokenReplacement(loopDefinition["sheet"], currentElement)

            if type == "findrow":
                if "column" not in loopDefinition:
                    raise ValueError(f"Missing 'column' in loop definition: {loopDefinition}")
                searchSlice = applyTokenReplacement(loopDefinition["column"], currentElement)
            else:
                if "row" not in loopDefinition:
                    raise ValueError(f"Missing 'row' in loop definition: {loopDefinition}")
                searchSlice = applyTokenReplacement(loopDefinition["row"], currentElement)
            
            data = [str(cell.value) if cell.value is not None else "" for cell in wb[sheet][searchSlice]]
            indices = [i for i, s in enumerate(data) if re.search(regex, s)]

            if len(indices) == 0:
                raise ValueError(f"Search regex '{regex}' not found in column '{searchSlice}' of sheet '{sheet}'")

            if "mode" not in loopDefinition or loopDefinition["mode"] == "first":
                indices = [indices[0]]
            elif loopDefinition["mode"] == "last":
                indices = [indices[-1]]
            elif loopDefinition["mode"] == "all":
                pass
            elif type(loopDefinition["mode"]) == int:
                indices = [indices[loopDefinition["mode"]]]
            else:
                raise ValueError(f"Invalid mode '{loopDefinition['mode']}' in definition: {loopDefinition}")

            offset = loopDefinition.get("offset", 0)
            indices = [i + offset + 1 for i in indices] # +1 to convert to 1-based index

            if type == "findrow":
                loopElements = indices
            else:
                loopElements = [get_column_letter(i) for i in indices]

        elif type == "looprows" or type == "loopcolumns":
            if "start" not in loopDefinition:
                raise ValueError(f"Missing 'start' in loop definition: {loopDefinition}")
            start = applyTokenReplacement(loopDefinition["start"], currentElement)
            if type == "loopcolumns":
                start = column_index_from_string(start)
            else:
                start = int(start)

            if "end" in loopDefinition and "count" in loopDefinition:
                raise ValueError("Cannot specify both 'end' and 'count' in loop definition")

            if "end" in loopDefinition:
                end = applyTokenReplacement(loopDefinition["end"], currentElement)
                if type == "loopcolumns":
                    end = column_index_from_string(end)
            elif "count" in loopDefinition:
                count = applyTokenReplacement(loopDefinition["count"], currentElement)
                end = start + count - 1
            elif "untilNoMatch" in loopDefinition and loopDefinition["untilNoMatch"]:
                raise ValueError("untilNoMatch is not implemented yet")
            else:
                raise ValueError("Must specify either 'end' or 'count' in loop definition")

            stride = loopDefinition.get("stride", 1)
            startOffset = loopDefinition.get("startOffset", 0)
            if startOffset != 0:
                start += startOffset
            endOffset = loopDefinition.get("endOffset", 0)
            if endOffset != 0:
                end += endOffset

            if end < start:
                raise ValueError(f"Start index {start} is greater than stop index {end} in definition: {loopDefinition}")

            indices = list(range(start, end + 1, stride))
            if type == "looprows":
                loopElements = indices
            else:
                loopElements = [get_column_letter(i) for i in indices]

        if len(loopDefinition) == 0:
            raise ValueError("Loop definition is empty")

        for i in range(len(loopElements)):
            copy = currentElement.copy()
            copy[token] = loopElements[i]
            resolveLookups(wb, elements, unprocessedDefinitions[1:], copy)
